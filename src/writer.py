# -*- coding: utf-8 -*-
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from validator import ValidationError

_UNIDADE_DESC: dict[str, str] = {
    "UN":  "UNIDADE",
    "KG":  "QUILOGRAMA",
    "G":   "GRAMA",
    "LT":  "LITRO",
    "ML":  "MILILITRO",
    "MT":  "METRO",
    "CM":  "CENTIMETRO",
    "CX":  "CAIXA",
    "PC":  "PECA",
    "PT":  "POTE",
    "SC":  "SACO",
    "DZ":  "DUZIA",
    "FD":  "FARDO",
    "BDJ": "BANDEJA",
}


def _build_grupo_subgrupo(df: pd.DataFrame) -> list[dict]:
    """
    Retorna lista única de grupos e subgrupos.
    Grupos primeiro (alfabético), depois subgrupos (alfabético).
    Coluna Grupo = 'GRUPO' ou 'SUBGRUPO'.
    Ordem = 0 para grupos, sequencial (1,2,3…) para subgrupos.
    """
    grupos: set[str] = set()
    subgrupos: set[str] = set()

    if "Grupo" in df.columns:
        for g in df["Grupo"].dropna():
            g = str(g).strip()
            if g:
                grupos.add(g)

    if "SubGrupo" in df.columns:
        for s in df["SubGrupo"].dropna():
            s = str(s).strip()
            if s:
                subgrupos.add(s)

    entries: list[dict] = []
    codigo = 1

    for g in sorted(grupos):
        entries.append({
            "Código": codigo, "Descrição": g, "Grupo": "GRUPO",
            "Loja": "", "Exibir na Venda": 1, "Ordem": 0, "Pedido Completo": 0,
        })
        codigo += 1

    for i, s in enumerate(sorted(subgrupos), start=1):
        entries.append({
            "Código": codigo, "Descrição": s, "Grupo": "SUBGRUPO",
            "Loja": "", "Exibir na Venda": 1, "Ordem": i, "Pedido Completo": 0,
        })
        codigo += 1

    return entries


def _write_grupo_subgrupo(wb: openpyxl.Workbook, df: pd.DataFrame) -> None:
    aba = next((n for n in wb.sheetnames if "grupo" in n.lower()), None)
    if not aba:
        return

    ws = wb[aba]

    # Detecta colunas pelo cabeçalho da linha 1
    col_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            col_map[str(val).strip()] = c

    # Limpa dados existentes (mantém cabeçalho)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    entries = _build_grupo_subgrupo(df)

    # Mapeamento nome-coluna → posição
    campo_col = {
        "Código":          col_map.get("Código") or col_map.get("C\xf3digo") or 1,
        "Descrição":       col_map.get("Descrição") or col_map.get("Descri\xe7\xe3o") or 2,
        "Grupo":           col_map.get("Grupo") or 3,
        "Loja":            col_map.get("Loja") or 4,
        "Exibir na Venda": col_map.get("Exibir na Venda") or 5,
        "Ordem":           col_map.get("Ordem") or 6,
        "Pedido Completo": col_map.get("Pedido Completo") or 7,
    }

    for i, e in enumerate(entries, start=2):
        for campo, col_idx in campo_col.items():
            ws.cell(i, col_idx, e.get(campo, ""))


def _write_unidades(wb: openpyxl.Workbook, df: pd.DataFrame) -> None:
    aba = next((n for n in wb.sheetnames if "unidade" in n.lower()), None)
    if not aba:
        return

    ws = wb[aba]

    # Detecta colunas pelo cabeçalho
    col_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            col_map[str(val).strip().lower()] = c

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    col_un  = col_map.get("unidade") or 1
    col_desc = col_map.get("descricao") or col_map.get("descrição") or 2
    col_loja = col_map.get("codigoloja") or 3

    if "Unidade" not in df.columns:
        return

    unidades = sorted({str(v).strip().upper() for v in df["Unidade"].dropna() if str(v).strip()})

    for i, un in enumerate(unidades, start=2):
        desc = _UNIDADE_DESC.get(un, un)
        ws.cell(i, col_un, un)
        ws.cell(i, col_desc, desc)
        ws.cell(i, col_loja, "")


def write_output(df: pd.DataFrame, output_path: Path, template_path: Path | None = None) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if template_path and template_path.exists():
        wb = openpyxl.load_workbook(template_path, keep_vba=True)

        # Aba Produtos
        ws = wb["Produtos"]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for row_data in dataframe_to_rows(df, index=False, header=False):
            ws.append(row_data)

        # Aba Grupo E Subgrupo
        _write_grupo_subgrupo(wb, df)

        # Aba Unidades
        _write_unidades(wb, df)

        wb.save(output_path)
        return

    # Fallback: gera xlsx simples (sem VBA)
    output_path = output_path.with_suffix(".xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Produtos")
        ws = writer.sheets["Produtos"]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center")

        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        ws.freeze_panes = "A2"


def write_error_report(errors: list[ValidationError], output_path: Path) -> None:
    if not errors:
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)

    df_err = pd.DataFrame([
        {
            "Linha (cliente)": e.row,
            "Campo":           e.field,
            "Valor Informado": e.value,
            "Motivo do Erro":  e.reason,
        }
        for e in errors
    ])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_err.to_excel(writer, index=False, sheet_name="Erros")
        ws = writer.sheets["Erros"]

        red = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        for cell in ws[1]:
            cell.fill = red
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = yellow

        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

        ws.freeze_panes = "A2"
